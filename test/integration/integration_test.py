import asyncio
from datetime import datetime, UTC
import json
import os
import sys
import contextlib
from unittest.mock import patch
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
import uvicorn
import websockets  # type: ignore[import]
from websockets.exceptions import ConnectionClosed  # type: ignore[import]
from openpyxl import Workbook, load_workbook  # type: ignore[import]
from openpyxl.styles import Font, Alignment  # type: ignore[import]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import]

# Ensure project root and app package are on sys.path so this can be run
# either from the repository root (python -m test.integration.task_reminder)
# or from within the test directory (python -m integration.task_reminder).
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../"))
if project_root not in sys.path:
    sys.path.insert(0, project_root)
app_path = os.path.join(project_root, "app")
if app_path not in sys.path:
    sys.path.insert(0, app_path)


# Load environment variables (including GOOGLE_API_KEY, DB settings, etc.) from the project .env
load_dotenv(os.path.join(project_root, ".env"))

# Override DB_* with LOCAL_DB_* so the app and cleanup both use the local database
os.environ["DB_HOST"] = os.environ["LOCAL_DB_HOST"]
os.environ["DB_PORT"] = os.environ.get("LOCAL_DB_PORT", "5432")
_local_db = os.environ.get("LOCAL_DB_NAME", "local").strip()
if _local_db.lower() == "postgres":
    raise SystemExit(
        '\nLOCAL_DB_NAME must not be "postgres" for integration tests (default admin DB, easy to confuse with app data).\n'
        "Run:  python test/setup_local_postgres.py\n"
        "Then in repo root .env set:  LOCAL_DB_NAME=local\n"
    )
os.environ["LOCAL_DB_NAME"] = _local_db
os.environ["DB_NAME"] = _local_db
os.environ["DB_USER"] = os.environ["LOCAL_DB_USER"]
os.environ["DB_PASSWORD"] = os.environ.get("LOCAL_DB_PASSWORD", "")

from app.main import app

# Freeze the simulated "current time" to a fixed point so tests are deterministic.
# All agent tool prompts will see this as the current date/time.
TEST_FROZEN_TIME = datetime(2026, 3, 1, 9, 0, 0, tzinfo=ZoneInfo("America/Los_Angeles"))

WS_HOST = "127.0.0.1"
WS_PORT = 8765
WS_URL_TEMPLATE = f"ws://{WS_HOST}:{WS_PORT}/ws/{{user_id}}"

# Excel file path for test inputs/outputs (static file, not created at runtime)
TEST_INPUTS_FILE = os.path.join(os.path.dirname(__file__), "test_inputs.xlsx")




def _read_test_inputs_from_excel() -> list[tuple[str, dict, str]]:
    """Read test input payloads from the Excel file.
    
    Returns:
        List of tuples (label, payload, setup_query) for each test row
    
    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If no test inputs are found in the file
    """
    if not os.path.exists(TEST_INPUTS_FILE):
        raise FileNotFoundError(
            f"Test inputs file not found: {TEST_INPUTS_FILE}\n"
            f"Please create the file first by running:\n"
            f"  python -m test.integration.create_test_inputs"
        )
    
    wb = load_workbook(TEST_INPUTS_FILE)
    ws: Worksheet = wb.active  # type: ignore[assignment]
    
    test_inputs = []
    # Skip header row (row 1), start from row 2
    for row in ws.iter_rows(min_row=2, values_only=False):
        if not row[0].value:  # Skip empty rows
            continue
        
        label = str(row[0].value)
        # Read input string from Input column (column B, index 1)
        input_text = str(row[1].value).strip() if row[1].value else ""
        if not input_text:
            print(f"Warning: Empty input for test '{label}', skipping...")
            continue
        
        # Read optional setup query from column E (index 4)
        setup_query = str(row[4].value).strip() if len(row) > 4 and row[4].value else ""

        # Convert string input to payload dict format for websocket
        # The websocket handler expects {"turns": {"message": "..."}} format
        payload = {"turns": {"message": input_text}}
        
        test_inputs.append((label, payload, setup_query))
    
    wb.close()
    
    if not test_inputs:
        raise ValueError(
            f"No valid test inputs found in {TEST_INPUTS_FILE}.\n"
            f"Please add test inputs to the file. Each row should have:\n"
            f"  - Test Label (column A)\n"
            f"  - Input (column B) with a plain text string"
        )
    
    print(f"Read {len(test_inputs)} test input(s) from {TEST_INPUTS_FILE}")
    return test_inputs


async def _run_single_session(user_id: str, label: str, payload: dict) -> tuple[str, dict, str, str]:
    """Open a websocket session, send one prompt then close.

    Waits up to MAX_WAIT_SECONDS for the agent to finish (end_conversation signal),
    then requests the scratchpad once.

    Returns:
        Tuple of (label, payload, transcription_output, scratchpad)
    """
    MAX_WAIT_SECONDS = 30

    ws_url = WS_URL_TEMPLATE.format(user_id=user_id)
    transcription_lines = []
    agent_text_parts = []

    scratchpad_entries = None
    done_event = asyncio.Event()       # fires when end_conversation received or timeout
    scratchpad_event = asyncio.Event() # fires when scratchpad message arrives

    async with websockets.connect(ws_url) as ws:

        async def reader():
            nonlocal scratchpad_entries, agent_text_parts
            try:
                async for raw in ws:
                    try:
                        msg = json.loads(raw)
                    except json.JSONDecodeError:
                        continue

                    if msg.get("type") == "scratchpad":
                        scratchpad_entries = msg.get("scratchpad", [])
                        scratchpad_event.set()
                        continue

                    if "audio" in msg:
                        continue

                    if "output_text" in msg:
                        agent_text_parts.append(msg["output_text"])
                    if "input_text" in msg:
                        if agent_text_parts:
                            transcription_lines.append(f"AGENT: {' '.join(agent_text_parts)}")
                            agent_text_parts = []
                        transcription_lines.append(f"USER: {msg['input_text']}")
                    if msg.get("end_conversation"):
                        if agent_text_parts:
                            transcription_lines.append(f"AGENT: {' '.join(agent_text_parts)}")
                            agent_text_parts = []
                        transcription_lines.append("END: Conversation ended by server")
                        done_event.set()
                        break
            except (asyncio.CancelledError, websockets.exceptions.ConnectionClosed):
                pass
            except Exception:
                pass

        reader_task = asyncio.create_task(reader())

        await ws.send(json.dumps(payload))

        # Wait up to MAX_WAIT_SECONDS for the agent to finish
        try:
            await asyncio.wait_for(done_event.wait(), timeout=MAX_WAIT_SECONDS)
        except asyncio.TimeoutError:
            print(f"  [{label}] Max wait reached ({MAX_WAIT_SECONDS}s), requesting scratchpad now.")

        # Request scratchpad once and wait up to 5s for the response
        print(f"  [{label}] Requesting scratchpad...")
        try:
            await ws.send(json.dumps({"type": "request_scratchpad"}))
            await asyncio.wait_for(scratchpad_event.wait(), timeout=5.0)
            print(f"  [{label}] Scratchpad received ({len(scratchpad_entries or [])} entries).")
        except asyncio.TimeoutError:
            print(f"  [{label}] Warning: no scratchpad response.")
        except Exception as e:
            print(f"  [{label}] Warning: could not request scratchpad: {e}")

        reader_task.cancel()
        with contextlib.suppress(asyncio.CancelledError):
            await reader_task

    if agent_text_parts:
        transcription_lines.append(f"AGENT: {' '.join(agent_text_parts)}")

    scratchpad_str = ""
    if scratchpad_entries:
        try:
            scratchpad_str = json.dumps(scratchpad_entries, indent=2)
        except Exception as e:
            print(f"Warning: Failed to format scratchpad: {e}")
            scratchpad_str = str(scratchpad_entries)

    transcription_output = "\n".join(transcription_lines)
    return (label, payload, transcription_output, scratchpad_str)


def _db_connect():
    """Return a new psycopg2 connection using the current DB_* env vars."""
    import psycopg2
    return psycopg2.connect(
        host=os.environ["DB_HOST"],
        port=os.environ.get("DB_PORT", "5432"),
        database=os.environ["DB_NAME"],
        user=os.environ["DB_USER"],
        password=os.environ.get("DB_PASSWORD", ""),
    )


def _ensure_db_connection() -> None:
    """Exit early with a clear message if local Postgres is not reachable."""
    import psycopg2
    host = os.environ["DB_HOST"]
    port = os.environ.get("DB_PORT", "5432")
    try:
        conn = _db_connect()
        conn.close()
    except psycopg2.OperationalError as e:
        raise SystemExit(
            f"\nCannot connect to PostgreSQL at {host}:{port}.\n"
            "This integration test snapshots and cleans DB tables; a running server is required.\n"
            "- Start PostgreSQL (Windows: Win+R → services.msc → start the PostgreSQL service).\n"
            "- Match .env LOCAL_DB_HOST / LOCAL_DB_PORT / LOCAL_DB_NAME / LOCAL_DB_USER / LOCAL_DB_PASSWORD.\n"
            "Create the schema:  python test/setup_local_postgres.py  (from repo root)\n\n"
            f"Underlying error: {e}\n"
        ) from e


def _ensure_app_schema() -> None:
    """Exit if LOCAL_DB_NAME points at a database without the app schema."""
    conn = _db_connect()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT 1 FROM information_schema.tables "
            "WHERE table_schema = 'public' AND table_name = 'sessions' LIMIT 1"
        )
        if cur.fetchone() is None:
            db = os.environ.get("DB_NAME", "?")
            raise SystemExit(
                f'\nDatabase "{db}" has no public.sessions table.\n'
                "Set LOCAL_DB_NAME in the repo root .env to the database where you ran "
                "psql ... -f schema.sql (usually local, not postgres).\n"
                f"Current value: LOCAL_DB_NAME={db!r}\n"
                "Create the schema:  python test/setup_local_postgres.py  (from repo root)\n"
            )
        cur.close()
    finally:
        conn.close()


def _run_db_query(query: str) -> None:
    """Execute a raw SQL statement against the local database."""
    conn = _db_connect()
    try:
        cursor = conn.cursor()
        cursor.execute(query)
        conn.commit()
        cursor.close()
    finally:
        conn.close()


def _snapshot_db() -> str:
    """Return a formatted string snapshot of all rows in the key tables."""
    tables = ["users", "sessions", "tasks", "messages", "pending_text_message_jobs"]
    sections = []
    conn = _db_connect()
    try:
        cursor = conn.cursor()
        for table in tables:
            try:
                cursor.execute(f"SELECT * FROM {table}")
                rows = cursor.fetchall()
                col_names = [desc[0] for desc in cursor.description] if cursor.description else []
                if not rows:
                    sections.append(f"[{table}]: (empty)")
                else:
                    header = " | ".join(col_names)
                    divider = "-" * len(header)
                    row_lines = [" | ".join(str(v) for v in row) for row in rows]
                    sections.append(f"[{table}]:\n{header}\n{divider}\n" + "\n".join(row_lines))
            except Exception as e:
                sections.append(f"[{table}]: ERROR - {e}")
        cursor.close()
    finally:
        conn.close()
    return "\n\n".join(sections)


def _cleanup_db() -> None:
    """Delete all rows from data tables after each test, leaving users intact."""
    tables = ["tasks", "messages", "sessions", "pending_text_message_jobs"]
    for table in tables:
        try:
            _run_db_query(f"DELETE FROM {table}")
            print(f"  Cleaned up table: {table}")
        except Exception as e:
            print(f"  Warning: could not clean table '{table}': {e}")


async def _websocket_client_test(user_id: str) -> list[tuple[str, dict, str, str, str, str]]:
    """Connect to the FastAPI websocket endpoint and exercise the full agent stack.

    Reads test inputs from Excel file and runs tests for each payload.
    Opens a fresh websocket (and thus Gemini) session for each prompt, running them serially.
    Runs the setup query from column E before each test, snapshots DB before and after,
    then cleans up all data tables after.
    
    Returns:
        List of tuples (label, payload, transcription_output, scratchpad, db_before, db_after)
        for each test session.
    """
    test_inputs = _read_test_inputs_from_excel()
    test_results = []

    # Run tests for each input from the Excel file
    for label, payload, setup_query in test_inputs:
        if setup_query:
            print(f"\nRunning setup query for '{label}'...")
            for statement in [s.strip() for s in setup_query.split(";") if s.strip()]:
                try:
                    _run_db_query(statement)
                except Exception as e:
                    print(f"  Warning: setup statement failed for '{label}': {e}")

        print(f"\nSnapshotting DB state before test '{label}'...")
        db_before = _snapshot_db()

        label_out, payload_out, transcription, scratchpad = await _run_single_session(
            user_id=user_id,
            label=label,
            payload=payload,
        )

        print(f"\nSnapshotting DB state after test '{label}'...")
        db_after = _snapshot_db()

        test_results.append((label_out, payload_out, transcription, scratchpad, db_before, db_after))

        print(f"\nCleaning up DB after test '{label}'...")
        _cleanup_db()

    return test_results


def _write_results_to_excel(test_results: list[tuple[str, dict, str, str, str, str]]) -> str:
    """Write test outputs back to the Excel file.
    
    Updates columns C–G for each test row:
      C - Transcription Output
      D - Scratchpad
      E - Setup Query (read-only, preserved)
      F - DB State Before
      G - DB State After
    
    Args:
        test_results: List of tuples (label, payload, transcription_output, scratchpad, db_before, db_after)
    
    Returns:
        Path to the Excel file
    """
    if not os.path.exists(TEST_INPUTS_FILE):
        raise FileNotFoundError(
            f"Test inputs file not found: {TEST_INPUTS_FILE}\n"
            f"Please create the file first by running:\n"
            f"  python -m test.integration.create_test_inputs"
        )
    
    wb = load_workbook(TEST_INPUTS_FILE)
    ws: Worksheet = wb.active  # type: ignore[assignment]
    
    # Build lookup maps keyed by label
    transcription_map = {label: transcription for label, _, transcription, _, _, _ in test_results}
    scratchpad_map    = {label: scratchpad    for label, _, _, scratchpad, _, _ in test_results}
    db_before_map     = {label: db_before     for label, _, _, _, db_before, _ in test_results}
    db_after_map      = {label: db_after      for label, _, _, _, _, db_after in test_results}

    # Ensure column F and G headers exist
    for col, header in [(6, "DB State Before"), (7, "DB State After")]:
        try:
            cell = ws.cell(row=1, column=col)
            if not cell.value:
                cell.value = header  # type: ignore[assignment]
                cell.font = Font(bold=True)
        except Exception:
            pass

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if not row[0].value:
            continue
        
        label = str(row[0].value)

        def _write_cell(col: int, value: str) -> None:
            try:
                cell = ws.cell(row=row_idx, column=col)
                cell.value = value  # type: ignore[assignment]
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            except Exception:
                pass

        if label in transcription_map:
            _write_cell(3, transcription_map[label])
        if label in scratchpad_map:
            _write_cell(4, scratchpad_map[label])
        if label in db_before_map:
            _write_cell(6, db_before_map[label])
        if label in db_after_map:
            _write_cell(7, db_after_map[label])

    # Column widths
    ws.column_dimensions["C"].width = 80   # Transcription Output
    ws.column_dimensions["D"].width = 80   # Scratchpad
    ws.column_dimensions["E"].width = 60   # Setup Query
    ws.column_dimensions["F"].width = 80   # DB State Before
    ws.column_dimensions["G"].width = 80   # DB State After
    
    wb.save(TEST_INPUTS_FILE)
    wb.close()
    return TEST_INPUTS_FILE


async def run_server_and_test() -> None:
    """Start uvicorn in-process and run the websocket client test against it.
    
    Freezes datetime.now() to TEST_FROZEN_TIME for the duration of the run so
    all agent prompts see a deterministic current date/time.
    """
    test_user_id = "2ba330c0-a999-46f8-ba2c-855880bdcf5b"

    frozen = TEST_FROZEN_TIME
    print(f"⏱  Frozen test time: {frozen.strftime('%A, %B %d, %Y at %I:%M %p (%Z)')}")
    _ensure_db_connection()
    _ensure_app_schema()

    def _frozen_now(tz=None):
        return frozen.astimezone(tz) if tz else frozen

    modules_to_patch = [
        "user_session_manager",
        "agents.tool_agents.create_tasks_tool_agent",
    ]

    with contextlib.ExitStack() as stack:
        for mod in modules_to_patch:
            mock_dt = stack.enter_context(patch(f"{mod}.datetime"))
            mock_dt.now.side_effect = _frozen_now
            mock_dt.fromisoformat = datetime.fromisoformat
            mock_dt.side_effect = lambda *a, **kw: datetime(*a, **kw)

        config = uvicorn.Config(app, host=WS_HOST, port=WS_PORT, log_level="info", ws="websockets")
        server = uvicorn.Server(config)

        server_task = asyncio.create_task(server.serve())
        # Give the server a moment to start
        await asyncio.sleep(1.5)

        try:
            test_results = await _websocket_client_test(user_id=test_user_id)
            excel_path = _write_results_to_excel(test_results)
            print(f"\n✓ Test results saved to Excel file: {excel_path}")
        except (FileNotFoundError, ValueError) as e:
            print(f"\n✗ Error: {e}")
            return
        finally:
            server.should_exit = True
            await server_task


def main() -> None:
    """
    Convenience entrypoint so you can run:

        python -m test.integration.task_reminder

    This will start the FastAPI/uvicorn server in-process, open a websocket
    connection to `/ws/{user_id}`, send a user message that should exercise `general_thinking_agent` and
    downstream agentic tools. All agent speech transcriptions are printed.
    """
    asyncio.run(run_server_and_test())


if __name__ == "__main__":
    main()

