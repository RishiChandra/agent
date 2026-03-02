import asyncio
from datetime import datetime, UTC
import json
import os
import sys
import contextlib

from dotenv import load_dotenv
import uvicorn
import websockets  # type: ignore[import]
from websockets.exceptions import ConnectionClosed  # type: ignore[import]
from openpyxl import Workbook, load_workbook  # type: ignore[import]
from openpyxl.styles import Font, Alignment  # type: ignore[import]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import]

# Ensure project root and app package are on sys.path so this can be run
# either from the repository root (python -m test.integration.task_reminder)
# or from within the test directory (python -m integration.task_reminder).
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../"))
if project_root not in sys.path:
    sys.path.insert(0, project_root)
app_path = os.path.join(project_root, "app")
if app_path not in sys.path:
    sys.path.insert(0, app_path)


# Load environment variables (including GOOGLE_API_KEY, DB settings, etc.) from the project .env
load_dotenv(os.path.join(project_root, ".env"))

from app.main import app

WS_HOST = "127.0.0.1"
WS_PORT = 8765
WS_URL_TEMPLATE = f"ws://{WS_HOST}:{WS_PORT}/ws/{{user_id}}"

# Excel file path for test inputs/outputs (static file, not created at runtime)
TEST_INPUTS_FILE = os.path.join(os.path.dirname(__file__), "test_inputs.xlsx")




def _read_test_inputs_from_excel() -> list[tuple[str, dict]]:
    """Read test input payloads from the Excel file.
    
    Returns:
        List of tuples (label, payload) for each test row
    
    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If no test inputs are found in the file
    """
    if not os.path.exists(TEST_INPUTS_FILE):
        raise FileNotFoundError(
            f"Test inputs file not found: {TEST_INPUTS_FILE}\n"
            f"Please create the file first by running:\n"
            f"  python -m test.integration.create_test_inputs"
        )
    
    wb = load_workbook(TEST_INPUTS_FILE)
    ws: Worksheet = wb.active  # type: ignore[assignment]
    
    test_inputs = []
    # Skip header row (row 1), start from row 2
    for row in ws.iter_rows(min_row=2, values_only=False):
        if not row[0].value:  # Skip empty rows
            continue
        
        label = str(row[0].value)
        # Read input string from Input column (column B, index 1)
        input_text = str(row[1].value).strip() if row[1].value else ""
        if not input_text:
            print(f"Warning: Empty input for test '{label}', skipping...")
            continue
        
        # Convert string input to payload dict format for websocket
        # The websocket handler expects {"turns": {"message": "..."}} format
        payload = {"turns": {"message": input_text}}
        
        test_inputs.append((label, payload))
    
    wb.close()
    
    if not test_inputs:
        raise ValueError(
            f"No valid test inputs found in {TEST_INPUTS_FILE}.\n"
            f"Please add test inputs to the file. Each row should have:\n"
            f"  - Test Label (column A)\n"
            f"  - Input (column B) with a plain text string"
        )
    
    print(f"Read {len(test_inputs)} test input(s) from {TEST_INPUTS_FILE}")
    return test_inputs


async def _run_single_session(user_id: str, label: str, payload: dict) -> tuple[str, dict, str, str]:
    """Open a websocket session, send one prompt then close.
    
    Returns:
        Tuple of (label, payload, transcription_output, scratchpad)
    """
    ws_url = WS_URL_TEMPLATE.format(user_id=user_id)
    transcription_lines = []
    agent_text_parts = []  # Collect all agent output chunks

    scratchpad_entries = None
    
    async with websockets.connect(ws_url) as ws:
        # Reader task: collects whatever websocket_handler sends (TranscriptionHandler outputs)
        async def reader():
            nonlocal scratchpad_entries, agent_text_parts
            try:
                async for raw in ws:
                    try:
                        msg = json.loads(raw)
                    except json.JSONDecodeError:
                        # Skip non-JSON messages
                        continue

                    # Capture scratchpad message sent by server before closing
                    if msg.get("type") == "scratchpad":
                        scratchpad_entries = msg.get("scratchpad", [])
                        continue

                    # Skip audio messages (they're just noise in logs)
                    if "audio" in msg:
                        continue

                    if "output_text" in msg:
                        # Collect agent text chunks instead of creating separate lines
                        agent_text_parts.append(msg['output_text'])
                    if "input_text" in msg:
                        # When we see user input, finalize any pending agent text
                        if agent_text_parts:
                            combined_agent_text = " ".join(agent_text_parts)
                            transcription_lines.append(f"AGENT: {combined_agent_text}")
                            agent_text_parts = []
                        line = f"USER: {msg['input_text']}"
                        transcription_lines.append(line)
                    if msg.get("end_conversation"):
                        # Finalize any pending agent text before ending
                        if agent_text_parts:
                            combined_agent_text = " ".join(agent_text_parts)
                            transcription_lines.append(f"AGENT: {combined_agent_text}")
                            agent_text_parts = []
                        line = "END: Conversation ended by server"
                        transcription_lines.append(line)
                        break
            except (asyncio.CancelledError, websockets.exceptions.ConnectionClosed):
                # Expected when connection closes or task is cancelled
                pass
            except Exception:
                # Silently ignore other errors
                pass

        reader_task = asyncio.create_task(reader())

        # Send the text payload
        await ws.send(json.dumps(payload))
        # Let the agent speak; tune this delay as needed
        await asyncio.sleep(30)

        # Request scratchpad from server before closing
        # Send a special message to request scratchpad
        try:
            await ws.send(json.dumps({"type": "request_scratchpad"}))
            # Wait a bit for the server to respond with scratchpad
            await asyncio.sleep(1.0)
        except Exception:
            pass  # Connection might already be closing
        
        reader_task.cancel()
        with contextlib.suppress(asyncio.CancelledError):
            await reader_task

    # Finalize any remaining agent text that wasn't followed by user input or end
    if agent_text_parts:
        combined_agent_text = " ".join(agent_text_parts)
        transcription_lines.append(f"AGENT: {combined_agent_text}")

    # Format scratchpad entries as a readable string
    scratchpad_str = ""
    if scratchpad_entries:
        try:
            # Format scratchpad entries as JSON for readability
            scratchpad_str = json.dumps(scratchpad_entries, indent=2)
        except Exception as e:
            print(f"Warning: Failed to format scratchpad: {e}")
            scratchpad_str = str(scratchpad_entries)
    else:
        scratchpad_str = ""

    transcription_output = "\n".join(transcription_lines)
    return (label, payload, transcription_output, scratchpad_str)


async def _websocket_client_test(user_id: str) -> list[tuple[str, dict, str, str]]:
    """Connect to the FastAPI websocket endpoint and exercise the full agent stack.

    Reads test inputs from Excel file and runs tests for each payload.
    Opens a fresh websocket (and thus Gemini) session for each prompt, running them serially.
    
    Returns:
        List of tuples (label, payload, transcription_output, scratchpad) for each test session.
    """
    test_inputs = _read_test_inputs_from_excel()
    test_results = []

    # Run tests for each input from the Excel file
    for label, payload in test_inputs:
        result = await _run_single_session(
            user_id=user_id,
            label=label,
            payload=payload,
        )
        test_results.append(result)

    return test_results


def _write_results_to_excel(test_results: list[tuple[str, dict, str, str]]) -> str:
    """Write transcription outputs and scratchpad back to the Excel file.
    
    Updates the Transcription Output column (column C) and Scratchpad column (column D) for each test row.
    
    Args:
        test_results: List of tuples (label, payload, transcription_output, scratchpad)
    
    Returns:
        Path to the Excel file
    """
    if not os.path.exists(TEST_INPUTS_FILE):
        raise FileNotFoundError(
            f"Test inputs file not found: {TEST_INPUTS_FILE}\n"
            f"Please create the file first by running:\n"
            f"  python -m test.integration.create_test_inputs"
        )
    
    wb = load_workbook(TEST_INPUTS_FILE)
    ws: Worksheet = wb.active  # type: ignore[assignment]
    
    # Create mappings of label to transcription output and scratchpad
    transcription_map = {label: transcription for label, _, transcription, _ in test_results}
    scratchpad_map = {label: scratchpad for label, _, _, scratchpad in test_results}
    
    # Update transcription outputs (column C, index 3) and scratchpad (column D, index 4)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if not row[0].value:  # Skip empty rows
            continue
        
        label = str(row[0].value)
        if label in transcription_map:
            # Update the Transcription Output column (column C, index 3)
            try:
                transcription_cell = ws.cell(row=row_idx, column=3)
                transcription_cell.value = transcription_map[label]  # type: ignore[assignment]
                transcription_cell.alignment = Alignment(wrap_text=True, vertical="top")
            except Exception:
                # If cell is merged, try to unmerge or access parent cell
                # For now, skip if there's an issue
                pass
        
        if label in scratchpad_map:
            # Update the Scratchpad column (column D, index 4)
            try:
                scratchpad_cell = ws.cell(row=row_idx, column=4)
                scratchpad_cell.value = scratchpad_map[label]  # type: ignore[assignment]
                scratchpad_cell.alignment = Alignment(wrap_text=True, vertical="top")
            except Exception:
                # If cell is merged, try to unmerge or access parent cell
                # For now, skip if there's an issue
                pass
    
    # Ensure column widths are set
    ws.column_dimensions["C"].width = 80  # Transcription Output
    ws.column_dimensions["D"].width = 80  # Scratchpad
    
    # Set header for scratchpad column if it doesn't exist
    try:
        header_cell = ws.cell(row=1, column=4)
        if header_cell.value is None:
            header_cell.value = "Scratchpad"  # type: ignore[assignment]
            header_cell.font = Font(bold=True)
    except Exception:
        # If cell is merged or inaccessible, skip header update
        pass
    
    wb.save(TEST_INPUTS_FILE)
    wb.close()
    return TEST_INPUTS_FILE


async def run_server_and_test() -> None:
    """Start uvicorn in-process and run the websocket client test against it."""
    test_user_id = "2ba330c0-a999-46f8-ba2c-855880bdcf5b"

    config = uvicorn.Config(app, host=WS_HOST, port=WS_PORT, log_level="info", ws="websockets")
    server = uvicorn.Server(config)

    server_task = asyncio.create_task(server.serve())
    # Give the server a moment to start
    await asyncio.sleep(1.5)

    try:
        test_results = await _websocket_client_test(user_id=test_user_id)
        excel_path = _write_results_to_excel(test_results)
        print(f"\n✓ Test results saved to Excel file: {excel_path}")
    except (FileNotFoundError, ValueError) as e:
        print(f"\n✗ Error: {e}")
        return
    finally:
        server.should_exit = True
        await server_task


def main() -> None:
    """
    Convenience entrypoint so you can run:

        python -m test.integration.task_reminder

    This will start the FastAPI/uvicorn server in-process, open a websocket
    connection to `/ws/{user_id}`, send a user message that should exercise `general_thinking_agent` and
    downstream agentic tools. All agent speech transcriptions are printed.
    """
    asyncio.run(run_server_and_test())


if __name__ == "__main__":
    main()

